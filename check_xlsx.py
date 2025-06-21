import pandas as pd
import openpyxl

# Load the Excel file
file_path = 'public/t001/annotated_transcript.xlsx'
wb = openpyxl.load_workbook(file_path)
print('Available sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    print(f'\n=== Sheet: {sheet_name} ===')
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    print('Columns:', list(df.columns))
    print('Shape (rows, cols):', df.shape)
    
    # Get feature columns (excluding #, speaker, dialogue)
    feature_cols = [col for col in df.columns if col not in ['#', 'speaker', 'dialogue']]
    print('Feature columns:', feature_cols)
    
    # Check unique values in feature columns
    print('Unique values in features:')
    for col in feature_cols[:3]:  # Just first 3 to avoid too much output
        unique_vals = sorted(df[col].dropna().unique())
        print(f'  {col}: {unique_vals}')
    
    # Show a few rows with non-null feature values
    print('Sample rows with annotations:')
    non_null_rows = df[df[feature_cols].notna().any(axis=1)]
    if not non_null_rows.empty:
        print(non_null_rows.head(2)[['#', 'speaker'] + feature_cols[:3]]) 