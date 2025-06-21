import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('expert_annotations.xlsx')
ws = wb['What Students Are Saying']

print("=== HIDDEN COLUMNS DATA ===")

# Check the first few rows of hidden Expert columns
for col_num in range(1, 4):  # Columns A, B, C (Expert 1, 2, 3 hidden)
    col_letter = openpyxl.utils.get_column_letter(col_num)
    header = ws.cell(row=1, column=col_num).value
    print(f"\n--- {header} (Column {col_letter} - HIDDEN) ---")
    
    for row in range(2, min(6, ws.max_row + 1)):
        cell_value = ws.cell(row=row, column=col_num).value
        if cell_value:
            print(f"Row {row}: {str(cell_value)[:200]}...")
        else:
            print(f"Row {row}: (empty)")

print("\n=== VISIBLE COLUMNS DATA ===")

# Check the visible columns for comparison
for col_num in range(12, 15):  # Columns L, M, N (Expert 1, 2, 3 visible)
    col_letter = openpyxl.utils.get_column_letter(col_num)
    header = ws.cell(row=1, column=col_num).value
    print(f"\n--- {header} (Column {col_letter} - VISIBLE) ---")
    
    for row in range(2, min(6, ws.max_row + 1)):
        cell_value = ws.cell(row=row, column=col_num).value
        if cell_value:
            print(f"Row {row}: {str(cell_value)[:200]}...")
        else:
            print(f"Row {row}: (empty)") 